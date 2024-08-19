from docx import Document
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def extract_headings(docpath):
    """Extracts headings and associated information from a Word document."""
    doc = Document(docpath)
    headings = []
    current_heading = None

    # Iterate through each paragraph in the document
    for para in doc.paragraphs:
        if para.style.name.startswith('Heading 1'):
            if current_heading:
                headings.append(current_heading)
            current_heading = {
                "University": para.text.strip(), 
                "Info": []
            }
        elif current_heading and para.text.strip():
            current_heading["Info"].append(para.text.strip())  

    # Add the last heading and its items to the list
    if current_heading:
        headings.append(current_heading)
    
    return headings

def prepare_excel_data(headings):
    """Prepares data for writing to Excel."""
    excel_data = []
    for entry in headings:
        heading = entry["University"]
        info = entry["Info"]
        
        names = []
        roles = []
        countries = []
        
        for item in info:
            parts = item.split(',')
            names.append(parts[0].strip() if len(parts) > 0 else "")
            roles.append(parts[1].strip() if len(parts) > 1 else "")
            countries.append(parts[-1].strip() if len(parts) > 2 else "")
        
        # Combine names, roles, and countries into single strings with new lines separating them
        names_str = "\n".join(names)
        roles_str = "\n".join(roles)
        countries_str = "\n".join(countries)

        excel_data.append([heading, names_str, roles_str, countries_str])
    
    return excel_data

def save_to_excel(excel_data, excel_file):
    """Saves the prepared data to an Excel file."""
    # Create a DataFrame with columns
    df = pd.DataFrame(excel_data, columns=["University", "Names", "Roles", "Countries"])
    df.to_excel(excel_file, index=False)
    
    # Enable text wrapping in Excel and adjust column width with a max width limit
    wb = load_workbook(excel_file)
    ws = wb.active
    max_width = 50  # Set a maximum column width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            if cell.value:
                # Calculate the appropriate width but limit to max_width
                current_width = ws.column_dimensions[cell.column_letter].width
                new_width = min(len(cell.value) + 2, max_width)
                ws.column_dimensions[cell.column_letter].width = max(current_width, new_width)
    
    wb.save(excel_file)
    print(f"Excel file has been saved to {excel_file}")

def main(docpath, excel_file):
    """Main function to process the Word document and save the data to Excel."""
    headings = extract_headings(docpath)
    excel_data = prepare_excel_data(headings)
    save_to_excel(excel_data, excel_file)

# Run the script
docpath = r'C:\Users\UNICK\OneDrive\Documents\KakaoTalk Downloads\새 Microsoft Word 문서.docx'
excel_file = r'Y:\excel\excel_file.xlsx'
main(docpath, excel_file)
